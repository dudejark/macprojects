import openpyxl
import xlsxwriter
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

#this segment of the program opens a new workbook KYC_2 in Python39 folder
workbook = xlsxwriter.Workbook('/Users/amandudeja/Desktop/Python39/KYC_2.xlsx')


loc = ("/Users/amandudeja/Desktop/Python39/KYC_Organizer.xlsx")

wb_obj = openpyxl.load_workbook(loc)

sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_r = sheet_obj.max_row
l = 0
segname = []

for k in range(7, max_r + 1):
 for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = k, column = i)

    if i == 1 and cell_obj.value == None:

        l = 0

        break
    elif 'Segment Name:' in str(cell_obj.value):

        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'S.No.')

        segname.append(cell_obj.value)
        break
    else:
        for n in range(1, l):
         worksheet.write(n, 0, n)

        worksheet.write(l-1, i, cell_obj.value)

 l = l + 1
workbook.close()
#Above segment makes a new excel file KYC_2 and creates sheets according to segment names in KYC_Organizer file, adds Sr.no. column

xfile = openpyxl.load_workbook('/Users/amandudeja/Desktop/Python39/KYC_2.xlsx')

sheet = xfile['Sheet1']
sheet['L1'] ='1 ) Photo Identity Proof has not been obtained (EC)'
sheet['M1'] ='2 ) Copy of ID Proof proof have not been verified with originals'
sheet['N1'] ='3 ) Proof of Address has not been not obtained (EC)'
sheet['O1'] ='4 ) Copy of Address proof have not been verified with originals'
sheet['P1'] ='5 ) Recent photograph for opening of account not been obtained.'
sheet['Q1'] ='6 ) A copy of PAN card/Aadhar Card/ Form No. 60/61 has not been obtained (EC)'
sheet['R1'] ='7 ) Documents (AOF) not produced for verification.'
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 20
sheet.column_dimensions['M'].width = 20
sheet.column_dimensions['N'].width = 20
sheet.column_dimensions['O'].width = 20
sheet.column_dimensions['P'].width = 20
sheet.column_dimensions['Q'].width = 20
sheet.column_dimensions['R'].width = 20


sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)
sheet['N1'].alignment = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
fred = openpyxl.worksheet.pagebreak.Break(15, 50)
#sheet.page_breaks.append(fred)

sheet = xfile['Sheet2']
sheet['L1'] ='1 ) Copy of Certificate of Incorporation duly certfied has not been obtained'
sheet['M1'] ='2 ) Certified copy of Memorandum and Articles of Association has not been obtained'
sheet['N1'] ='3 ) Copy of PAN of the Company has not been obtained'
sheet['O1'] ='4 ) A resolution from the Board of Directors authorising its managers, officers or employees to transact on behalf of the company has not been obtained'
sheet['P1'] ='5 ) An officially valid document in respect of managers, officers or employees authorised under Board Resolution to transact on behalf of the company has not been obtained'
sheet['Q1'] ='6 ) Proof of current address of the Company has not been obtained'

sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 20
sheet.column_dimensions['M'].width = 20
sheet.column_dimensions['N'].width = 20
sheet.column_dimensions['O'].width = 20
sheet.column_dimensions['P'].width = 20
sheet.column_dimensions['Q'].width = 20

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['H1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
fred = openpyxl.worksheet.pagebreak.Break(5, 14)
#sheet.page_breaks.append(fred)

sheet = xfile['Sheet3']
sheet['L1'] ='1 ) "Certified copy of the following documents have not been obtained, Registration Certificate; Partnership deed An officially valid document of Proof of Identity, Proof of Address, in respect of the persons authorised to transact on behalf of the firm'
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 30

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['J1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
fred = openpyxl.worksheet.pagebreak.Break(5, 14)
#sheet.page_breaks.append(fred)

sheet = xfile['Sheet4']
sheet['L1'] ='1 ) In addition to the OVD applicable to the proprietor, any of the two following documents in the name of the proprietorship concern have not been obtained: Registration certificate , Certificate/license issued by the Municipal authorities under Shop & Establishment Act, Sales and income tax returns, CST/VAT certificate, Certificate/registration document issued by Sales Tax/Service Tax/Professional Tax authorities, IEC issued by the office of DGFT,Licence/Certificate of practice issued in the name of the proprietary concern by any professional body incorporated under statue Complete Income Tax Return in the name of the sole proprietor where the firm   s income is reflected, duly authenticated/ acknowledged by the Income Tax authorities. Utility bills such as electricity, water and landline telephone bills in the name of the proprietary concern.'
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 50

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['H1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
fred = openpyxl.worksheet.pagebreak.Break(5, 14)
#sheet.page_breaks.append(fred)

sheet = xfile['Sheet8']
sheet['K1'] ='1) KYC verification of all the office bearers of the SHG has not been carried out'
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 25

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['H1'].alignment = Alignment(wrap_text=True)
sheet['I1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
fred = openpyxl.worksheet.pagebreak.Break(5, 14)
#sheet.page_breaks.append(fred)


sheet = xfile['Sheet9']
sheet['K1'] ='1 ) Holder of Basic Savings Bank Deposit Account (BSBDA) is also maintaining other regular Savings Bank Deposit account in the Bank. Other existing Savings Bank deposit account (Non-BSBDA) not closed after 30 days from the date of opening a Basic Savings Bank Deposit Account. (EC-RBI-TIII)'
sheet['L1'] ='2 ) "i) In respect of Small Accounts, conditions stipulated are not strictly adhered - a. Balance at any point of time does not exceed Rs.50,000/- b.Aggregate of all credits in a year do not exceed Rs.1.00 lacc.Aggregate of all withdrawals and transfers in a month does not exceed Rs.10 thousand (EC-RBI-TIII) "'
sheet['M1'] ='3 ) ii) If any account is rendered ineligible for being classified as a small account due to credits / balance in the account exceeding the permissible limits, withdrawals are not allowed within the limit prescribed (Aggregate of all withdrawals and transfers in a month should not exceed Rs.10 thousand) (EC-RBI-TIII)'
sheet['N1'] ='4 ) iii) BSBD Accounts (PMJDY accounts are akin to BSBDAs) which are KYC compliant accounts are not treated as Small Accounts and are not subjected to the limitations applicable to such accounts (EC-RBI-TIII)'
sheet['O1'] ='5 ) Small Account not converted to BSBDA/other regular Savings Bank Account after 12/24 months upon submission of valid KYC documents.'
sheet['P1'] ='6 ) Further transactions allowed even after 24 months from the date of opening of Small Accounts without converting to regular Savings Bank Account'

sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 20
sheet.column_dimensions['L'].width = 20
sheet.column_dimensions['M'].width = 20
sheet.column_dimensions['N'].width = 20
sheet.column_dimensions['O'].width = 20
sheet.column_dimensions['P'].width = 20


sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['I1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
fred = openpyxl.worksheet.pagebreak.Break(5, 14)
#sheet.page_breaks.append(fred)

sheet = xfile['Sheet6']
sheet['K1'] ='1 ) "Certified copies of the following documents have not been obtained, Declaration from the Karta. Proof of Identification of Karta. Prescribed Joint Hindu Family Letter (COS 38) signed by all the adult co-parceners "'
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 25

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['k1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
fred = openpyxl.worksheet.pagebreak.Break(5, 14)
#sheet.page_breaks.append(fred)

sheet = xfile['Sheet5']
sheet['L1'] ='1 ) Trust Related'
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 25

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['H1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
fred = openpyxl.worksheet.pagebreak.Break(5, 14)
#sheet.page_breaks.append(fred)

sheet = xfile['Sheet7']
sheet['L1'] ='1. Reg: NRE A/cs'
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['k'].width = 12
sheet.column_dimensions['L'].width = 25

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['H1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
fred = openpyxl.worksheet.pagebreak.Break(5, 14)
#sheet.page_breaks.append(fred)

#print('At the end')
xfile.save('/Users/amandudeja/Desktop/Python39/KYC_2.xlsx')

workbook = xlsxwriter.Workbook('/Users/amandudeja/Desktop/Python39/SUMMARY_KYC.xlsx')
worksheet = workbook.add_worksheet('Summary')

xfile = openpyxl.load_workbook('/Users/amandudeja/Desktop/Python39/KYC_2.xlsx')
worksheet.write(0, 0, 'Summary')
worksheet.write(0, 1, 'Records')
total = 0
for k in range(1, 10):
    m = len(segname[k-1])
    segname1 = segname[k-1][14:m]
    sheetname = 'Sheet' + str(k)
    sheet = xfile[sheetname]
    xfile.active = k
    max_r = sheet.max_row-1
    worksheet.write(k, 0, segname1)
    worksheet.write(k, 1, max_r)
    total = total + max_r

worksheet.write(24, 0, 'Total Records')
worksheet.write(24, 1, total)


workbook.close()
xfile = openpyxl.load_workbook('/Users/amandudeja/Desktop/Python39/SUMMARY_KYC.xlsx')
sheet = xfile['Summary']
sheet.column_dimensions['A'].width = 50
for k in range(1, 26):
    sheet.cell(row=k, column=1).border = thin_border
    sheet.cell(row=k, column=2).border = thin_border
xfile.save('/Users/amandudeja/Desktop/Python39/SUMMARY_KYC.xlsx')
