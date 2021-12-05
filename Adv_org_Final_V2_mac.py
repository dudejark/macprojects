#thisispython


import openpyxl
import xlsxwriter

workbook = xlsxwriter.Workbook('/Users/amandudeja/Desktop/Python39/ADV_2.xlsx')


loc = ("/Users/amandudeja/Desktop/Python39/ADV_Organizer.xlsx")

wb_obj = openpyxl.load_workbook(loc)

sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_r = sheet_obj.max_row
l = 0
segname = []
prodcod9 = ""
for k in range(7, max_r + 1):
 for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = k, column = i)

    if i == 1 and cell_obj.value == None:

        l = 0

        break
    elif 'Segment Name:' in str(cell_obj.value):

        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'S.No.')

        segname.append(cell_obj.value)
        break
    else:
        for n in range(1, l):
         worksheet.write(n, 0, n)

        worksheet.write(l-1, i, cell_obj.value)

 l = l + 1
workbook.close()
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
xfile = openpyxl.load_workbook('/Users/amandudeja/Desktop/Python39/ADV_2.xlsx')

sheet = xfile['Sheet1']
sheet['L1'] ='PROD CODE'
sheet['M1'] ='PROD SHORT NAME'
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 33
sheet.column_dimensions['D'].width = 38
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 12
sheet.column_dimensions['M'].width = 12

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
  if k > 1:
    prodcod9  = sheet.cell(row=k, column=4).value[-9:]
    sheet.cell(row=k, column=12).value = prodcod9[:8]
fred = openpyxl.worksheet.pagebreak.Break(15, 14)
##sheet.page_breaks.append(fred)


sheet = xfile['Sheet2']
sheet['L1'] ='PROD CODE'
sheet['M1'] ='PROD SHORT NAME'
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 33
sheet.column_dimensions['D'].width = 38
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 12
sheet.column_dimensions['M'].width = 12

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
  if k > 1:
    prodcod9  = sheet.cell(row=k, column=4).value[-9:]
    sheet.cell(row=k, column=12).value = prodcod9[:8]
fred = openpyxl.worksheet.pagebreak.Break(15, 14)
##sheet.page_breaks.append(fred)

sheet = xfile['Sheet3']
sheet['L1'] ='PROD CODE'
sheet['M1'] ='PROD SHORT NAME'
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 33
sheet.column_dimensions['D'].width = 38
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 12
sheet.column_dimensions['M'].width = 12

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
  if k > 1:
    #prodcod9  = sheet.cell(row=k, column=4).value[-9:]
    sheet.cell(row=k, column=12).value = prodcod9[:8]
fred = openpyxl.worksheet.pagebreak.Break(15, 14)
##sheet.page_breaks.append(fred)

sheet = xfile['Sheet4']
sheet['L1'] ='PROD CODE'
sheet['M1'] ='PROD SHORT NAME'
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 33
sheet.column_dimensions['D'].width = 38
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 12
sheet.column_dimensions['M'].width = 12

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
  if k > 1:
    prodcod9  = sheet.cell(row=k, column=4).value[-9:]
    sheet.cell(row=k, column=12).value = prodcod9[:8]
fred = openpyxl.worksheet.pagebreak.Break(15, 14)
##sheet.page_breaks.append(fred)

"""sheet = xfile['Sheet5']
sheet['L1'] ='PROD CODE'
sheet['M1'] ='PROD SHORT NAME'
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 33
sheet.column_dimensions['D'].width = 38
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 12
sheet.column_dimensions['M'].width = 12

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
  if k > 1:
    prodcod9  = sheet.cell(row=k, column=4).value[-9:]
    sheet.cell(row=k, column=12).value = prodcod9[:8]
fred = openpyxl.worksheet.pagebreak.Break(15, 14)
##sheet.page_breaks.append(fred)"""


xfile.save('/Users/amandudeja/Desktop/Python39/ADV_2.xlsx')

workbook = xlsxwriter.Workbook('/Users/amandudeja/Desktop/Python39/SUMMARY_ADV.xlsx')
worksheet = workbook.add_worksheet('Summary')

xfile = openpyxl.load_workbook('/Users/amandudeja/Desktop/Python39/ADV_2.xlsx')
worksheet.write(0, 0, 'Summary')
worksheet.write(0, 1, 'Records')
total = 0
for k in range(1, 5):
    m = len(segname[k-1])
    segname1 = segname[k-1][14:m]
    sheetname = 'Sheet' + str(k)
    sheet = xfile[sheetname]
    xfile.active = k
    max_r = sheet.max_row-1
    worksheet.write(k, 0, segname1)
    worksheet.write(k, 1, max_r)
    total = total + max_r

worksheet.write(24, 0, 'Total Records')
worksheet.write(24, 1, total)


workbook.close()
xfile = openpyxl.load_workbook('/Users/amandudeja/Desktop/Python39/SUMMARY_ADV.xlsx')
sheet = xfile['Summary']
sheet.column_dimensions['A'].width = 60
for k in range(1, 26):
    sheet.cell(row=k, column=1).border = thin_border
    sheet.cell(row=k, column=2).border = thin_border
xfile.save('/Users/amandudeja/Desktop/Python39/SUMMARY_ADV.xlsx')

import pandas as pd
xl = pd.ExcelFile("/Users/amandudeja/Desktop/Python39/ADV_2.xlsx")
df1 = xl.parse("Sheet1")
df2 = xl.parse("Sheet2")
df3 = xl.parse("Sheet3")
df4 = xl.parse("Sheet4")
#df5 = xl.parse("Sheet5")

xl1 = pd.ExcelFile("/Users/amandudeja/Desktop/Python39/Product_codes.xlsx")
writer = pd.ExcelWriter('/Users/amandudeja/Desktop/Python39/ADV_3.xlsx')
df = xl1.parse("Sheet1")
results=df1.merge(df, on='PROD CODE', how='left')
results = results.sort_values(by=["PROD SHORT NAME_y","Account Number"])
results.to_excel(writer,sheet_name='Sheet1',columns=["S.No.",	"Account Number",	"Customer Name",	"Product Name (Code)",	"Segment",	"Sanction Date",	"Intrest Rate",	"Limit in Rs.",	"DP in Rs.",	"Outstanding in Rs.",	"New IRAC",	"PROD CODE",	"PROD SHORT NAME_y"
],na_rep='NA',index=False)
results=df2.merge(df, on='PROD CODE', how='left')
results = results.sort_values(by=["PROD SHORT NAME_y","Account Number"])
results.to_excel(writer,sheet_name='Sheet2',columns=["S.No.",	"Account Number",	"Customer Name",	"Product Name (Code)",	"Segment",	"Sanction Date",	"Intrest Rate",	"Limit in Rs.",	"DP in Rs.",	"Outstanding in Rs.",	"New IRAC",	"PROD CODE",	"PROD SHORT NAME_y"
],na_rep='NA',index=False)
results=df3.merge(df, on='PROD CODE', how='left')
results = results.sort_values(by=["PROD SHORT NAME_y","Account Number"])
results.to_excel(writer,sheet_name='Sheet3',columns=["S.No.",	"Account Number",	"Customer Name",	"Product Name (Code)",	"Segment",	"Sanction Date",	"Intrest Rate",	"Limit in Rs.",	"DP in Rs.",	"Outstanding in Rs.",	"New IRAC",	"PROD CODE",	"PROD SHORT NAME_y"
],na_rep='NA',index=False)
results=df4.merge(df, on='PROD CODE', how='left')
results = results.sort_values(by=["PROD SHORT NAME_y","Account Number"])
results.to_excel(writer,sheet_name='Sheet4',columns=["S.No.",	"Account Number",	"Customer Name",	"Product Name (Code)",	"Segment",	"Sanction Date",	"Intrest Rate",	"Limit in Rs.",	"DP in Rs.",	"Outstanding in Rs.",	"New IRAC",	"PROD CODE",	"PROD SHORT NAME_y"
],na_rep='NA',index=False)
"""results=df5.merge(df, on='PROD CODE', how='left')
results = results.sort_values(by=["PROD SHORT NAME_y","Account Number"])
results.to_excel(writer,sheet_name='Sheet5',columns=["S.No.",	"Account Number",	"Customer Name",	"Product Name (Code)",	"Segment",	"Sanction Date",	"Intrest Rate",	"Limit in Rs.",	"DP in Rs.",	"Outstanding in Rs.",	"New IRAC",	"PROD CODE",	"PROD SHORT NAME_y"
],na_rep='NA',index=False)"""
writer.save()

xfile = openpyxl.load_workbook('/Users/amandudeja/Desktop/Python39/ADV_3.xlsx')

sheet = xfile['Sheet1']
sheet['L1'] ='PROD CODE'
sheet['M1'] ='PROD SHORT NAME'
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 33
sheet.column_dimensions['D'].width = 38
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 12
sheet.column_dimensions['M'].width = 12

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['A1'].alignment = Alignment(wrap_text=True)
sheet['B1'].alignment = Alignment(wrap_text=True)
sheet['C1'].alignment = Alignment(wrap_text=True)
sheet['D1'].alignment = Alignment(wrap_text=True)
sheet['E1'].alignment = Alignment(wrap_text=True)
sheet['F1'].alignment = Alignment(wrap_text=True)
sheet['G1'].alignment = Alignment(wrap_text=True)
sheet['H1'].alignment = Alignment(wrap_text=True)
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  #sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
  if k>1:
   sheet.cell(row=k, column=1).value = k-1
 fred = openpyxl.worksheet.pagebreak.Break(15, 14)
##sheet.page_breaks.append(fred)

sheet = xfile['Sheet2']
sheet['L1'] ='PROD CODE'
sheet['M1'] ='PROD SHORT NAME'
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 33
sheet.column_dimensions['D'].width = 38
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 12
sheet.column_dimensions['M'].width = 12

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['A1'].alignment = Alignment(wrap_text=True)
sheet['B1'].alignment = Alignment(wrap_text=True)
sheet['C1'].alignment = Alignment(wrap_text=True)
sheet['D1'].alignment = Alignment(wrap_text=True)
sheet['E1'].alignment = Alignment(wrap_text=True)
sheet['F1'].alignment = Alignment(wrap_text=True)
sheet['G1'].alignment = Alignment(wrap_text=True)
sheet['H1'].alignment = Alignment(wrap_text=True)
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  #sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
  if k>1:
   sheet.cell(row=k, column=1).value = k-1
 fred = openpyxl.worksheet.pagebreak.Break(15, 14)
##sheet.page_breaks.append(fred)

sheet = xfile['Sheet3']
sheet['L1'] ='PROD CODE'
sheet['M1'] ='PROD SHORT NAME'
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 33
sheet.column_dimensions['D'].width = 38
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 12
sheet.column_dimensions['M'].width = 12

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['A1'].alignment = Alignment(wrap_text=True)
sheet['B1'].alignment = Alignment(wrap_text=True)
sheet['C1'].alignment = Alignment(wrap_text=True)
sheet['D1'].alignment = Alignment(wrap_text=True)
sheet['E1'].alignment = Alignment(wrap_text=True)
sheet['F1'].alignment = Alignment(wrap_text=True)
sheet['G1'].alignment = Alignment(wrap_text=True)
sheet['H1'].alignment = Alignment(wrap_text=True)
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  #sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
  if k>1:
   sheet.cell(row=k, column=1).value = k-1
 fred = openpyxl.worksheet.pagebreak.Break(15, 14)
##sheet.page_breaks.append(fred)

sheet = xfile['Sheet4']
sheet['L1'] ='PROD CODE'
sheet['M1'] ='PROD SHORT NAME'
sheet.column_dimensions['A'].width = 6
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 33
sheet.column_dimensions['D'].width = 38
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 12
sheet.column_dimensions['M'].width = 12

sheet.row_dimensions[1].height = 110
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight = 1
sheet.page_setup.fitToWidth = 1
sheet['A1'].alignment = Alignment(wrap_text=True)
sheet['B1'].alignment = Alignment(wrap_text=True)
sheet['C1'].alignment = Alignment(wrap_text=True)
sheet['D1'].alignment = Alignment(wrap_text=True)
sheet['E1'].alignment = Alignment(wrap_text=True)
sheet['F1'].alignment = Alignment(wrap_text=True)
sheet['G1'].alignment = Alignment(wrap_text=True)
sheet['H1'].alignment = Alignment(wrap_text=True)
sheet['I1'].alignment = Alignment(wrap_text=True)
sheet['J1'].alignment = Alignment(wrap_text=True)
sheet['K1'].alignment = Alignment(wrap_text=True)
sheet['L1'].alignment = Alignment(wrap_text=True)
sheet['M1'].alignment = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
max_col = sheet.max_column
max_r = sheet.max_row
for k in range(1, max_r + 1):
 for i in range(1, max_col + 1):
  sheet.cell(row=k, column=i).border = thin_border
  #sheet.cell(row=k, column=i).alignment = Alignment(wrap_text=True)
  if k>1:
   sheet.cell(row=k, column=1).value = k-1
 fred = openpyxl.worksheet.pagebreak.Break(15, 14)
##sheet.page_breaks.append(fred)


xfile.save('/Users/amandudeja/Desktop/Python39/ADV_3.xlsx')